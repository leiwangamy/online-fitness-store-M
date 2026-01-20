"""
Views for seller application and dashboard
"""
from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Sum, Q, Max as models_Max
from django.db import models as db_models
from django.utils import timezone
from datetime import timedelta, datetime, date
from decimal import Decimal
import csv
import json
from django.http import HttpResponse, JsonResponse

from .models import Seller, SellerMembershipPlan
from .forms import SellerApplicationForm, SellerProductForm
from .decorators import seller_required
from products.models import Product
from orders.models import OrderItem, Order, Refund


def seller_portal(request):
    """
    Seller portal landing page - first page users see.
    - If not logged in: Show login form with "Become a Seller" link
    - If logged in but no seller account: Show "Become a Seller" link
    - If logged in with seller account: Redirect to appropriate page
    """
    # If user is authenticated, check seller status
    if request.user.is_authenticated:
        try:
            seller = request.user.seller
            # If seller is approved, redirect to dashboard
            if seller.status == Seller.STATUS_APPROVED:
                return redirect('sellers:dashboard')
            # Otherwise, redirect to application status
            return redirect('sellers:application_status')
        except AttributeError:
            # User is logged in but has no seller account
            # Show the portal page with "Become a Seller" link
            pass
    
    # Not logged in or logged in without seller account
    return render(request, 'sellers/portal.html', {
        'is_authenticated': request.user.is_authenticated,
    })


def apply(request):
    """
    Page for users to apply to become a seller.
    - If user is logged in: Show only seller application form
    - If user is not logged in: Show signup form with email/password + seller application
    """
    # Check if user already has a seller profile
    if request.user.is_authenticated:
        try:
            seller = request.user.seller
            return redirect('sellers:application_status')
        except AttributeError:
            pass  # No seller profile yet, continue with application
    
    if request.method == 'POST':
        if request.user.is_authenticated:
            # Logged in user - just seller application form
            form = SellerApplicationForm(request.POST)
            if form.is_valid():
                seller = form.save(commit=False)
                seller.user = request.user
                seller.status = Seller.STATUS_PENDING
                seller.save()
                
                messages.success(
                    request,
                    "Your seller application has been submitted! "
                    "We'll review it and notify you once it's approved."
                )
                return redirect('sellers:application_status')
        else:
            # Not logged in - combined signup + seller application form
            from .forms import SellerSignupApplicationForm
            from allauth.account.models import EmailAddress
            form = SellerSignupApplicationForm(request.POST)
            if form.is_valid():
                # This will create user account and seller application
                user = form.save(request)
                
                # Check if email verification is required
                try:
                    email_address = EmailAddress.objects.get_for_user(user, user.email)
                    if not email_address.verified:
                        # Email verification required - redirect to email confirmation page
                        messages.success(
                            request,
                            "Your account has been created and seller application submitted! "
                            "Please check your email to verify your account. "
                            "Once verified, you can log in and check your application status."
                        )
                        return redirect('account_email_verification_sent')
                    else:
                        # Email already verified (shouldn't happen with mandatory verification, but just in case)
                        messages.success(
                            request,
                            "Your seller application has been submitted! "
                            "We'll review it and notify you once it's approved."
                        )
                        return redirect('sellers:application_status')
                except Exception:
                    # Fallback if EmailAddress doesn't exist yet
                    messages.success(
                        request,
                        "Your account has been created and seller application submitted! "
                        "Please check your email to verify your account."
                    )
                    return redirect('account_email_verification_sent')
    else:
        if request.user.is_authenticated:
            form = SellerApplicationForm()
        else:
            from .forms import SellerSignupApplicationForm
            form = SellerSignupApplicationForm()
    
    return render(request, 'sellers/apply.html', {
        'form': form,
        'is_authenticated': request.user.is_authenticated,
    })


def application_status(request):
    """
    Show the status of the user's seller application.
    If not logged in, show a message about email verification.
    Approved sellers are automatically redirected to the dashboard.
    """
    if not request.user.is_authenticated:
        messages.info(
            request,
            "Please log in to check your seller application status. "
            "If you just signed up, please verify your email first."
        )
        return redirect('account_login')
    
    try:
        seller = request.user.seller
    except AttributeError:
        # No seller profile - redirect to apply
        messages.info(request, "You haven't applied to become a seller yet.")
        return redirect('sellers:apply')
    
    # If seller is approved, redirect to dashboard
    if seller.status == Seller.STATUS_APPROVED:
        return redirect('sellers:dashboard')
    
    return render(request, 'sellers/application_status.html', {
        'seller': seller,
    })


@seller_required
def dashboard(request):
    """
    Seller dashboard showing overview stats.
    """
    seller = request.user.seller
    
    # Get seller's products
    products = Product.objects.filter(seller=seller)
    total_products = products.count()
    active_products = products.filter(is_active=True).count()
    
    # Get order statistics
    order_items = OrderItem.objects.filter(seller=seller).select_related('order')
    total_orders = order_items.values('order').distinct().count()
    total_earnings = order_items.aggregate(
        total=Sum('seller_earnings')
    )['total'] or Decimal('0.00')
    total_platform_fees = order_items.aggregate(
        total=Sum('platform_fee')
    )['total'] or Decimal('0.00')
    
    # Calculate available vs pending earnings based on fixed 7-day hold period
    # Orders become available 7 days after order creation (fixed period)
    hold_days = 7  # Fixed 7-day hold period for all sellers
    cutoff_date = timezone.now() - timedelta(days=hold_days)
    
    # Available: Earnings from orders older than hold period
    available_items = order_items.filter(order__created_at__lt=cutoff_date)
    available_earnings = available_items.aggregate(
        total=Sum('seller_earnings')
    )['total'] or Decimal('0.00')
    
    # Pending: Earnings from orders within hold period
    pending_items = order_items.filter(order__created_at__gte=cutoff_date)
    pending_earnings = pending_items.aggregate(
        total=Sum('seller_earnings')
    )['total'] or Decimal('0.00')
    
    # Next payout date: Date when oldest pending order becomes available
    next_payout_date = None
    oldest_pending_order = pending_items.order_by('order__created_at').first()
    if oldest_pending_order and oldest_pending_order.order:
        next_payout_date = oldest_pending_order.order.created_at + timedelta(days=hold_days)
    
    # Recent orders (last 5)
    recent_orders = order_items.select_related('order', 'product').order_by('-order__created_at')[:5]
    
    # Low stock products (quantity < 10)
    low_stock_products = products.filter(
        quantity_in_stock__lt=10,
        quantity_in_stock__gt=0,
        is_active=True
    )[:5]
    
    context = {
        'seller': seller,
        'total_products': total_products,
        'active_products': active_products,
        'total_orders': total_orders,
        'total_earnings': total_earnings,
        'total_platform_fees': total_platform_fees,
        'available_earnings': available_earnings,
        'pending_earnings': pending_earnings,
        'next_payout_date': next_payout_date,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
    }
    
    return render(request, 'sellers/dashboard.html', context)


@seller_required
def product_list(request):
    """
    List all products belonging to the seller.
    """
    seller = request.user.seller
    products = Product.objects.filter(seller=seller).select_related('category').order_by('-id')
    
    # Search functionality
    search_query = request.GET.get('q', '').strip()
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'sellers/product_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    })


@seller_required
def product_add(request):
    """
    Add a new product (seller-only) with multiple media items.
    """
    seller = request.user.seller
    
    if request.method == 'POST':
        from .forms import SellerProductForm
        from .formsets import ProductImageFormSet, ProductVideoFormSet, ProductAudioFormSet
        from products.models import Product
        
        form = SellerProductForm(request.POST, request.FILES)
        image_formset = ProductImageFormSet(request.POST, request.FILES, prefix='images')
        video_formset = ProductVideoFormSet(request.POST, request.FILES, prefix='videos')
        audio_formset = ProductAudioFormSet(request.POST, request.FILES, prefix='audios')
        
        if form.is_valid():
            product = form.save(commit=False)
            product.seller = seller  # Auto-set to logged-in seller
            product.is_featured = False  # Sellers can't feature products
            product.save()
            
            # Save formsets
            if image_formset.is_valid():
                image_formset.instance = product
                image_formset.save()
                # Clean up any empty image records (no file) - this handles cases where empty forms were saved
                from products.models import ProductImage
                ProductImage.objects.filter(product=product).filter(
                    db_models.Q(image__isnull=True) | db_models.Q(image='')
                ).delete()
                # Ensure at least one image is marked as main if images exist
                images = product.images.exclude(
                    db_models.Q(image__isnull=True) | db_models.Q(image='')
                )
                if images.exists() and not images.filter(is_main=True).exists():
                    first_image = images.first()
                    if first_image.image and first_image.image.name:
                        first_image.is_main = True
                        first_image.save()
            
            if video_formset.is_valid():
                video_formset.instance = product
                video_formset.save()
            
            if audio_formset.is_valid():
                audio_formset.instance = product
                audio_formset.save()
            
            # Check if formsets are valid
            if image_formset.is_valid() and video_formset.is_valid() and audio_formset.is_valid():
                messages.success(request, f"Product '{product.name}' has been added successfully!")
                return redirect('sellers:product_list')
    else:
        from .forms import SellerProductForm
        from .formsets import ProductImageFormSet, ProductVideoFormSet, ProductAudioFormSet
        from products.models import Product
        
        form = SellerProductForm()
        # Create empty formsets for new product
        image_formset = ProductImageFormSet(prefix='images')
        video_formset = ProductVideoFormSet(prefix='videos')
        audio_formset = ProductAudioFormSet(prefix='audios')
    
    return render(request, 'sellers/product_add.html', {
        'form': form,
        'image_formset': image_formset,
        'video_formset': video_formset,
        'audio_formset': audio_formset,
    })


@seller_required
def product_edit(request, product_id):
    """
    Edit a product (only if it belongs to the seller) with multiple media items.
    """
    seller = request.user.seller
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        from .forms import SellerProductForm
        from .formsets import ProductImageFormSet, ProductVideoFormSet, ProductAudioFormSet
        
        form = SellerProductForm(request.POST, request.FILES, instance=product)
        image_formset = ProductImageFormSet(request.POST, request.FILES, instance=product, prefix='images')
        video_formset = ProductVideoFormSet(request.POST, request.FILES, instance=product, prefix='videos')
        audio_formset = ProductAudioFormSet(request.POST, request.FILES, instance=product, prefix='audios')
        
        if form.is_valid():
            # Ensure seller can't change ownership or featured status
            product = form.save(commit=False)
            product.seller = seller  # Ensure ownership
            product.is_featured = False  # Sellers can't feature products
            product.save()
            
            # Save formsets
            if image_formset.is_valid():
                image_formset.save()
                # Clean up any empty image records (no file) - this handles cases where empty forms were saved
                from products.models import ProductImage
                ProductImage.objects.filter(product=product).filter(
                    db_models.Q(image__isnull=True) | db_models.Q(image='')
                ).delete()
                # Ensure at least one image is marked as main if images exist
                images = product.images.exclude(
                    db_models.Q(image__isnull=True) | db_models.Q(image='')
                )
                if images.exists() and not images.filter(is_main=True).exists():
                    first_image = images.first()
                    if first_image.image and first_image.image.name:
                        first_image.is_main = True
                        first_image.save()
            
            if video_formset.is_valid():
                video_formset.save()
            
            if audio_formset.is_valid():
                audio_formset.save()
            
            # Check if formsets are valid
            if image_formset.is_valid() and video_formset.is_valid() and audio_formset.is_valid():
                messages.success(request, f"Product '{product.name}' has been updated successfully!")
                return redirect('sellers:product_list')
    else:
        from .forms import SellerProductForm
        from .formsets import ProductImageFormSet, ProductVideoFormSet, ProductAudioFormSet
        
        form = SellerProductForm(instance=product)
        image_formset = ProductImageFormSet(instance=product, prefix='images')
        video_formset = ProductVideoFormSet(instance=product, prefix='videos')
        audio_formset = ProductAudioFormSet(instance=product, prefix='audios')
    
    return render(request, 'sellers/product_edit.html', {
        'form': form,
        'product': product,
        'image_formset': image_formset,
        'video_formset': video_formset,
        'audio_formset': audio_formset,
    })


@seller_required
def product_delete(request, product_id):
    """
    Delete a product (only if it belongs to the seller).
    """
    seller = request.user.seller
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        messages.success(request, f"Product '{product_name}' has been deleted.")
        return redirect('sellers:product_list')
    
    return render(request, 'sellers/product_delete.html', {
        'product': product,
    })


@seller_required
def order_list(request):
    """
    List all orders containing seller's products.
    """
    seller = request.user.seller
    
    # Get all orders that contain this seller's products
    order_items = OrderItem.objects.filter(seller=seller).select_related('order', 'order__user', 'product')
    
    # Get unique orders
    order_ids = order_items.values_list('order_id', flat=True).distinct()
    orders = Order.objects.filter(id__in=order_ids).select_related('user', 'pickup_location').prefetch_related('items').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'sellers/order_list.html', {
        'seller': seller,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'status_choices': Order.STATUS_CHOICES,
    })


@seller_required
def order_detail(request, order_id):
    """
    View order details with shipping address and allow status updates.
    """
    seller = request.user.seller
    
    # Get order and verify it contains seller's products
    order = get_object_or_404(Order, id=order_id)
    order_items = OrderItem.objects.filter(order=order, seller=seller)
    
    if not order_items.exists():
        messages.error(request, "You don't have permission to view this order.")
        return redirect('sellers:order_list')
    
    # Get all items in this order (seller's items only)
    seller_items = order_items.select_related('product')
    
    # Calculate seller-specific totals
    seller_subtotal = sum(item.line_total for item in seller_items)
    seller_earnings = sum(item.seller_earnings for item in seller_items)
    seller_platform_fees = sum(item.platform_fee for item in seller_items)
    
    # Handle status update
    if request.method == 'POST':
        new_status = request.POST.get('status')
        tracking_number = request.POST.get('tracking_number', '').strip()
        shipping_carrier = request.POST.get('shipping_carrier', '').strip()
        
        if new_status in dict(Order.STATUS_CHOICES):
            # Only allow sellers to update to certain statuses
            allowed_statuses = [Order.STATUS_PROCESSING, Order.STATUS_SHIPPED]
            if new_status in allowed_statuses:
                order.status = new_status
                if tracking_number:
                    order.tracking_number = tracking_number
                if shipping_carrier:
                    order.shipping_carrier = shipping_carrier
                order.save()
                messages.success(request, f"Order #{order.id} status updated to {order.get_status_display()}.")
            else:
                messages.error(request, "You can only update order status to 'Processing' or 'Shipped'.")
        else:
            messages.error(request, "Invalid status selected.")
        
        return redirect('sellers:order_detail', order_id=order.id)
    
    # Import refund policy functions
    from services.refund_policy import is_within_refund_window, can_seller_auto_refund
    
    # Get refunds for this seller's items
    seller_refunds = Refund.objects.filter(
        order=order,
        seller=seller
    ).select_related('order_item', 'created_by').order_by('-created_at')
    
    # Attach refund eligibility directly to each item (Option A - cleanest approach)
    for item in seller_items:
        # Check if item can be refunded
        can_refund = (
            order.status in [Order.STATUS_PAID, Order.STATUS_PROCESSING, Order.STATUS_SHIPPED] and
            is_within_refund_window(order) and
            not Refund.objects.filter(
                order_item=item,
                status__in=[Refund.STATUS_SUCCEEDED, Refund.STATUS_PROCESSING]
            ).exists()
        )
        
        if can_refund:
            can_auto_refund = can_seller_auto_refund(order, seller, is_partial=False, has_dispute=None)
            item.refund_allowed = True
            item.refund_mode = "auto" if can_auto_refund else "request"
            item.refund_reason = None
        else:
            item.refund_allowed = False
            item.refund_mode = None
            # Determine reason
            if order.status not in [Order.STATUS_PAID, Order.STATUS_PROCESSING, Order.STATUS_SHIPPED]:
                item.refund_reason = f"Order status is {order.get_status_display()}"
            elif not is_within_refund_window(order):
                item.refund_reason = "Outside 7-day refund window"
            elif Refund.objects.filter(
                order_item=item,
                status__in=[Refund.STATUS_SUCCEEDED, Refund.STATUS_PROCESSING]
            ).exists():
                item.refund_reason = "Already refunded or processing"
            else:
                item.refund_reason = "Not eligible for refund"
    
    return render(request, 'sellers/order_detail.html', {
        'seller': seller,
        'order': order,
        'order_items': seller_items,
        'seller_subtotal': seller_subtotal,
        'seller_earnings': seller_earnings,
        'seller_platform_fees': seller_platform_fees,
        'seller_refunds': seller_refunds,
        'status_choices': Order.STATUS_CHOICES,
        'carrier_choices': Order.CARRIER_CHOICES,
    })


@seller_required
def refund_order_item(request, order_item_id):
    """
    Seller-initiated refund for an order item.
    If qualifies for auto-refund -> process immediately via Stripe
    Else -> create refund request for admin approval
    """
    from django.db import transaction
    from django.http import JsonResponse, HttpResponseForbidden
    from services.refund_policy import can_seller_auto_refund
    from services.stripe_refunds import create_stripe_refund, StripeRefundError, _to_cents
    
    if not request.user.is_authenticated:
        return HttpResponseForbidden()
    
    seller = request.user.seller
    order_item = get_object_or_404(OrderItem, id=order_item_id)
    
    # Verify seller owns this order item
    if order_item.seller_id != seller.id:
        return JsonResponse({"ok": False, "error": "You don't have permission to refund this item."}, status=403)
    
    order = order_item.order
    
    # Check if order has payment intent ID
    if not order.payment_intent_id:
        return JsonResponse({"ok": False, "error": "Missing Stripe payment reference."}, status=400)
    
    # Check if already refunded
    existing_refunds = Refund.objects.filter(
        order=order,
        order_item=order_item,
        status__in=[Refund.STATUS_SUCCEEDED, Refund.STATUS_PROCESSING]
    )
    if existing_refunds.exists():
        return JsonResponse({"ok": False, "error": "This item has already been refunded."}, status=400)
    
    # Determine refund amount (full item amount)
    refund_amount = order_item.line_total  # Refund the full line total to customer
    is_partial = False  # For now, we only support full item refunds
    
    # Check if seller can auto-refund
    auto = can_seller_auto_refund(order, seller, is_partial=is_partial, has_dispute=None)
    
    try:
        with transaction.atomic():
            # Create refund record
            refund = Refund.objects.create(
                order=order,
                seller=seller,
                order_item=order_item,
                amount=refund_amount,
                reason=request.POST.get('reason', 'Seller initiated refund'),
                created_by=request.user,
                status=Refund.STATUS_APPROVED if auto else Refund.STATUS_REQUESTED,
            )
            
            if not auto:
                # Notify admin (you can add email/notification here)
                messages.info(request, "Refund request submitted for admin review.")
                return JsonResponse({
                    "ok": True,
                    "status": "requested",
                    "message": "Refund request submitted for admin review."
                })
            
            # Auto refund path - process immediately
            refund.status = Refund.STATUS_PROCESSING
            refund.save(update_fields=["status"])
            
            # Process Stripe refund
            from services.stripe_refunds import _to_cents
            stripe_refund_id = create_stripe_refund(
                payment_intent_id=order.payment_intent_id,
                amount_cents=_to_cents(refund_amount),
                reason="requested_by_customer"
            )
            
            refund.stripe_refund_id = stripe_refund_id
            refund.status = Refund.STATUS_SUCCEEDED
            refund.save(update_fields=["stripe_refund_id", "status"])
            
            # Update order status
            # Check if all items in order are refunded
            total_refunded = Refund.objects.filter(
                order=order,
                status=Refund.STATUS_SUCCEEDED
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            if total_refunded >= order.total:
                order.status = Order.STATUS_REFUNDED
            else:
                order.status = Order.STATUS_PARTIALLY_REFUNDED
            order.save(update_fields=["status"])
            
            messages.success(request, f"Refund of ${refund_amount} processed successfully.")
            return JsonResponse({
                "ok": True,
                "status": "succeeded",
                "message": "Refund processed successfully."
            })
            
    except StripeRefundError as e:
        refund.status = Refund.STATUS_FAILED
        refund.save(update_fields=["status"])
        messages.error(request, f"Refund failed: {str(e)}")
        return JsonResponse({"ok": False, "status": "failed", "error": str(e)}, status=400)
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@seller_required
def earnings_statement(request):
    """
    View earnings statement with detailed breakdown, date filtering, and transaction log.
    """
    from datetime import datetime, date
    import csv
    from django.http import HttpResponse
    
    seller = request.user.seller
    now = timezone.now()
    today = now.date()
    
    # Handle date filter
    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Calculate date range based on period or custom dates
    if period == 'today':
        start_date = today
        end_date = today
    elif period == 'last_7_days':
        start_date = today - timedelta(days=6)
        end_date = today
    elif period == 'this_month':
        start_date = date(today.year, today.month, 1)
        end_date = today
    elif period == 'last_month':
        import calendar
        if today.month == 1:
            start_date = date(today.year - 1, 12, 1)
            end_date = date(today.year - 1, 12, 31)
        else:
            start_date = date(today.year, today.month - 1, 1)
            last_day = calendar.monthrange(today.year, today.month - 1)[1]
            end_date = date(today.year, today.month - 1, last_day)
    elif period == 'custom' and date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except:
            start_date = date(today.year, today.month, 1)
            end_date = today
            period = 'this_month'
    else:
        start_date = date(today.year, today.month, 1)
        end_date = today
        period = 'this_month'
    
    # Convert to datetime for filtering
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    # Get order items within date range
    order_items = OrderItem.objects.filter(
        seller=seller,
        order__created_at__gte=start_datetime,
        order__created_at__lte=end_datetime
    ).select_related('order', 'product', 'order__user').order_by('order__created_at')
    
    # Get refunds within date range
    refunds = Refund.objects.filter(
        seller=seller,
        status=Refund.STATUS_SUCCEEDED,
        created_at__gte=start_datetime,
        created_at__lte=end_datetime
    ).select_related('order', 'order_item').order_by('created_at')
    
    # Build transaction log in bank-style format
    transactions = []
    running_balance = Decimal('0.00')
    
    # Track tax by type (Products vs Memberships)
    tax_products_gst = Decimal('0.00')
    tax_products_pst = Decimal('0.00')
    tax_memberships_gst = Decimal('0.00')
    tax_memberships_pst = Decimal('0.00')
    
    # Helper function to check if order item is membership
    def is_membership_order(item):
        """Check if order item is a membership"""
        # Check if product name contains membership keywords or if it's from seller membership plan
        product_name_lower = item.product.name.lower()
        if 'membership' in product_name_lower or 'subscription' in product_name_lower:
            return True
        # Could also check product category or type if available
        return False
    
    # Process order items
    for item in order_items:
        is_membership = is_membership_order(item)
        source = "Membership" if is_membership else "Product"
        
        # Calculate taxes
        gst = item.line_total * Decimal('0.05') if item.product.charge_gst else Decimal('0.00')
        pst = item.line_total * Decimal('0.07') if item.product.charge_pst else Decimal('0.00')
        
        # Track tax by type
        if is_membership:
            tax_memberships_gst += gst
            tax_memberships_pst += pst
        else:
            tax_products_gst += gst
            tax_products_pst += pst
        
        # 1. Product/Membership earnings transaction (positive)
        running_balance += item.seller_earnings
        transactions.append({
            'date': item.order.created_at,
            'source': source,
            'description': f"Order #{item.order.id} – {item.product.name}",
            'amount': item.seller_earnings,
            'balance': running_balance,
            'order_id': item.order.id,
            'type': 'order',
            'is_membership': is_membership,
            'gst': gst,
            'pst': pst,
        })
        
        # 2. Commission fee transaction (negative, if commission exists)
        if item.platform_fee > Decimal('0.00'):
            running_balance -= item.platform_fee
            transactions.append({
                'date': item.order.created_at,
                'source': 'Commission',
                'description': f"Platform fee ({seller.commission_rate * 100:.0f}%)",
                'amount': -item.platform_fee,  # Negative amount
                'balance': running_balance,
                'order_id': item.order.id,
                'type': 'commission',
                'is_membership': is_membership,
                'gst': Decimal('0.00'),
                'pst': Decimal('0.00'),
            })
    
    # Process refunds
    for refund in refunds:
        # Determine if refund is for membership
        is_membership = False
        if refund.order_item:
            is_membership = is_membership_order(refund.order_item)
        
        source = "Membership" if is_membership else "Product"
        refund_description = f"Order #{refund.order.id} refund"
        if refund.order_item:
            refund_description = f"Order #{refund.order.id} – {refund.order_item.product.name} refund"
        if refund.reason:
            refund_description = f"{refund_description} ({refund.reason})"
        
        # 1. Refund transaction (negative)
        running_balance -= refund.amount
        transactions.append({
            'date': refund.created_at,
            'source': 'Refund',
            'description': refund_description,
            'amount': -refund.amount,  # Negative amount
            'balance': running_balance,
            'order_id': refund.order.id,
            'type': 'refund',
            'is_membership': is_membership,
            'gst': Decimal('0.00'),
            'pst': Decimal('0.00'),
        })
        
        # 2. Commission reversal (positive, if original order had commission)
        if refund.order_item and refund.order_item.platform_fee > Decimal('0.00'):
            # Calculate proportional commission reversal
            # If full refund, reverse full commission; if partial, calculate proportion
            original_amount = refund.order_item.line_total
            if original_amount > Decimal('0.00'):
                commission_reversal = (refund.amount / original_amount) * refund.order_item.platform_fee
                running_balance += commission_reversal
                transactions.append({
                    'date': refund.created_at,
                    'source': 'Commission',
                    'description': 'Commission reversal',
                    'amount': commission_reversal,
                    'balance': running_balance,
                    'order_id': refund.order.id,
                    'type': 'commission_reversal',
                    'is_membership': is_membership,
                    'gst': Decimal('0.00'),
                    'pst': Decimal('0.00'),
                })
    
    # Sort by date (oldest first for statement)
    transactions.sort(key=lambda x: x['date'])
    
    # Calculate period totals
    # TOTAL REVENUE: All positive amounts (orders + commission reversals)
    total_revenue = sum(t['amount'] for t in transactions if t['amount'] > 0)
    # TOTAL COMMISSION: All negative commission amounts (absolute value)
    total_commission = abs(sum(t['amount'] for t in transactions if t['amount'] < 0 and t['type'] == 'commission'))
    # NET CHANGE: Sum of all transaction amounts
    net_change = sum(t['amount'] for t in transactions)
    
    # Calculate tax totals
    total_gst = tax_products_gst + tax_memberships_gst
    total_pst = tax_products_pst + tax_memberships_pst
    total_tax = total_gst + total_pst
    
    # Handle CSV export
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="earnings_statement_{start_date}_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Earnings Statement', f'{start_date} to {end_date}'])
        writer.writerow([])
        writer.writerow(['Date', 'Source', 'Description', 'Amount', 'Balance'])
        
        for t in transactions:
            amount_str = f"+{t['amount']:.2f}" if t['amount'] >= 0 else f"{t['amount']:.2f}"
            writer.writerow([
                t['date'].strftime('%Y-%m-%d'),
                t['source'],
                t['description'],
                amount_str,
                f"${t['balance']:.2f}"
            ])
        
        writer.writerow([])
        writer.writerow(['Total Gross Revenue', '', '', f"+${total_revenue:.2f}", ''])
        writer.writerow(['Platform Commission', '', '', f"-${total_commission:.2f}", ''])
        writer.writerow(['', '', '', '---', ''])
        writer.writerow(['Net Change', '', '', f"${net_change:.2f}", ''])
        writer.writerow(['Ending Balance', '', '', '', f"${running_balance:.2f}"])
        writer.writerow([])
        writer.writerow(['Tax Summary (Reference Only)'])
        writer.writerow(['Products'])
        writer.writerow(['  GST', f"${tax_products_gst:.2f}"])
        writer.writerow(['  PST', f"${tax_products_pst:.2f}"])
        writer.writerow(['Memberships'])
        writer.writerow(['  GST', f"${tax_memberships_gst:.2f}"])
        writer.writerow(['  PST', f"${tax_memberships_pst:.2f}"])
        writer.writerow(['Total Tax Collected'])
        writer.writerow(['  GST', f"${total_gst:.2f}"])
        writer.writerow(['  PST', f"${total_pst:.2f}"])
        writer.writerow(['  Total', f"${total_tax:.2f}"])
        
        return response
    
    context = {
        'seller': seller,
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'start_date': start_date,
        'end_date': end_date,
        'transactions': transactions,
        'total_revenue': total_revenue,
        'total_commission': total_commission,
        'net_change': net_change,
        'final_balance': running_balance,
        'tax_products_gst': tax_products_gst,
        'tax_products_pst': tax_products_pst,
        'tax_memberships_gst': tax_memberships_gst,
        'tax_memberships_pst': tax_memberships_pst,
        'total_gst': total_gst,
        'total_pst': total_pst,
        'total_tax': total_tax,
    }
    
    return render(request, 'sellers/earnings_statement.html', context)


@seller_required
def data_export(request):
    """
    Seller data export page with filters and export options.
    Allows exporting Orders, Products, Refunds, and Statement.
    """
    seller = request.user.seller
    
    # Get filter parameters
    export_type = request.GET.get('export_type', '')
    export_format = request.GET.get('export_format', 'csv')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    status_filter = request.GET.get('status', '')
    product_filter = request.GET.get('product', '')
    
    # Get seller's products for dropdown
    seller_products = Product.objects.filter(seller=seller).order_by('name')
    
    # Handle export
    if export_type and request.GET.get('generate') == 'true':
        if export_type == 'orders':
            return export_orders(seller, start_date, end_date, status_filter, product_filter, export_format)
        elif export_type == 'products':
            return export_products(seller, status_filter, export_format)
        elif export_type == 'refunds':
            return export_refunds(seller, start_date, end_date, status_filter, export_format)
        elif export_type == 'statement':
            return export_statement(seller, start_date, end_date, export_format)
        else:
            messages.error(request, "Invalid export type.")
            return redirect('sellers:data_export')
    
    # Get order status choices
    order_status_choices = Order.STATUS_CHOICES
    
    # Check which export formats are available
    excel_available = False
    pdf_available = False
    try:
        import openpyxl
        excel_available = True
    except ImportError:
        pass
    
    try:
        import reportlab
        pdf_available = True
    except ImportError:
        pass
    
    return render(request, 'sellers/data_export.html', {
        'seller': seller,
        'export_type': export_type,
        'export_format': export_format,
        'start_date': start_date,
        'end_date': end_date,
        'status_filter': status_filter,
        'product_filter': product_filter,
        'seller_products': seller_products,
        'order_status_choices': order_status_choices,
        'excel_available': excel_available,
        'pdf_available': pdf_available,
    })


def export_orders(seller, start_date, end_date, status_filter, product_filter, export_format='csv'):
    """Export seller's orders in the specified format"""
    if export_format == 'csv':
        return export_orders_csv(seller, start_date, end_date, status_filter, product_filter)
    elif export_format == 'excel':
        return export_orders_excel(seller, start_date, end_date, status_filter, product_filter)
    elif export_format == 'json':
        return export_orders_json(seller, start_date, end_date, status_filter, product_filter)
    else:
        return export_orders_csv(seller, start_date, end_date, status_filter, product_filter)


def export_orders_csv(seller, start_date, end_date, status_filter, product_filter):
    """Export seller's orders to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f"orders_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Header row
    writer.writerow([
        'Order ID', 'Order Date', 'Customer Email', 'Order Status',
        'Product Name', 'Quantity', 'Unit Price',
        'Line Total', 'Platform Fee', 'Seller Earnings',
        'Shipping Address', 'Tracking Number', 'Shipping Carrier'
    ])
    
    # Get order items
    order_items = OrderItem.objects.filter(seller=seller).select_related(
        'order', 'order__user', 'product'
    )
    
    # Apply filters
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            order_items = order_items.filter(order__created_at__gte=start_dt)
        except:
            pass
    
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(
                datetime.strptime(end_date, '%Y-%m-%d').date(),
                datetime.max.time()
            ))
            order_items = order_items.filter(order__created_at__lte=end_dt)
        except:
            pass
    
    if status_filter:
        order_items = order_items.filter(order__status=status_filter)
    
    if product_filter:
        order_items = order_items.filter(product_id=product_filter)
    
    # Write data rows
    for item in order_items.order_by('-order__created_at'):
        order = item.order
        product = item.product
        
        # Build shipping address string
        shipping_address = ""
        if order.is_pickup and order.pickup_location:
            shipping_address = f"PICKUP: {order.pickup_location.name}, {order.pickup_location.address1}, {order.pickup_location.city}"
        else:
            parts = []
            if order.ship_name:
                parts.append(order.ship_name)
            if order.ship_address1:
                parts.append(order.ship_address1)
            if order.ship_city:
                parts.append(order.ship_city)
            if order.ship_province:
                parts.append(order.ship_province)
            if order.ship_postal_code:
                parts.append(order.ship_postal_code)
            shipping_address = ", ".join(parts)
        
        writer.writerow([
            order.id,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.user.email if order.user else 'Guest',
            order.get_status_display(),
            product.name,
            item.quantity,
            f"{item.price:.2f}",
            f"{item.line_total:.2f}",
            f"{item.platform_fee:.2f}",
            f"{item.seller_earnings:.2f}",
            shipping_address,
            order.tracking_number or '',
            order.get_shipping_carrier_display() if order.shipping_carrier else '',
        ])
    
    return response


def export_products_csv(seller, status_filter):
    """Export seller's products to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Header row
    writer.writerow([
        'Product ID', 'Name', 'Category', 'Price',
        'Quantity in Stock', 'Is Active', 'Is Digital', 'Is Service',
        'Is Featured', 'Created At', 'Updated At'
    ])
    
    # Get products
    products = Product.objects.filter(seller=seller).select_related('category')
    
    # Apply status filter (is_active)
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    
    # Write data rows
    for product in products.order_by('-created_at'):
        writer.writerow([
            product.id,
            product.name,
            product.category.name if product.category else '',
            f"{product.price:.2f}",
            product.quantity_in_stock,
            'Yes' if product.is_active else 'No',
            'Yes' if product.is_digital else 'No',
            'Yes' if product.is_service else 'No',
            'Yes' if product.is_featured else 'No',
            product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '',
            product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else '',
        ])
    
    return response


def export_refunds(seller, start_date, end_date, status_filter, export_format='csv'):
    """Export seller's refunds in the specified format"""
    if export_format == 'csv':
        return export_refunds_csv(seller, start_date, end_date, status_filter)
    elif export_format == 'excel':
        try:
            return export_refunds_excel(seller, start_date, end_date, status_filter)
        except ImportError:
            return export_refunds_csv(seller, start_date, end_date, status_filter)
    elif export_format == 'json':
        return export_refunds_json(seller, start_date, end_date, status_filter)
    else:
        return export_refunds_csv(seller, start_date, end_date, status_filter)


def export_refunds_csv(seller, start_date, end_date, status_filter):
    """Export seller's refunds to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f"refunds_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Header row
    writer.writerow([
        'Refund ID', 'Order ID', 'Product Name', 'Amount',
        'Reason', 'Status', 'Created By', 'Created At',
        'Stripe Refund ID'
    ])
    
    # Get refunds
    refunds = Refund.objects.filter(seller=seller).select_related(
        'order', 'order_item', 'order_item__product', 'created_by'
    )
    
    # Apply filters
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            refunds = refunds.filter(created_at__gte=start_dt)
        except:
            pass
    
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(
                datetime.strptime(end_date, '%Y-%m-%d').date(),
                datetime.max.time()
            ))
            refunds = refunds.filter(created_at__lte=end_dt)
        except:
            pass
    
    if status_filter:
        refunds = refunds.filter(status=status_filter)
    
    # Write data rows
    for refund in refunds.order_by('-created_at'):
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        elif refund.order_item:
            product_name = 'N/A'
        else:
            product_name = 'Full Order Refund'
        
        writer.writerow([
            refund.id,
            refund.order.id,
            product_name,
            f"{refund.amount:.2f}",
            refund.reason or '',
            refund.get_status_display(),
            refund.created_by.email if refund.created_by else '',
            refund.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            refund.stripe_refund_id or '',
        ])
    
    return response


def export_statement(seller, start_date, end_date, export_format='csv'):
    """Export seller's earnings statement in the specified format"""
    if export_format == 'csv':
        return export_statement_csv(seller, start_date, end_date)
    elif export_format == 'excel':
        try:
            return export_statement_excel(seller, start_date, end_date)
        except ImportError:
            return export_statement_csv(seller, start_date, end_date)
    elif export_format == 'json':
        return export_statement_json(seller, start_date, end_date)
    elif export_format == 'pdf':
        try:
            return export_statement_pdf(seller, start_date, end_date)
        except ImportError:
            return export_statement_csv(seller, start_date, end_date)
    else:
        return export_statement_csv(seller, start_date, end_date)


def export_statement_csv(seller, start_date, end_date):
    """Export seller's earnings statement to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f"statement_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Calculate date range
    now = timezone.now()
    today = now.date()
    
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            start_dt = date(today.year, today.month, 1)
            end_dt = today
    else:
        start_dt = date(today.year, today.month, 1)
        end_dt = today
    
    start_datetime = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_dt, datetime.max.time()))
    
    # Get order items within date range
    order_items = OrderItem.objects.filter(
        seller=seller,
        order__created_at__gte=start_datetime,
        order__created_at__lte=end_datetime
    ).select_related('order', 'product', 'order__user').order_by('order__created_at')
    
    # Build transaction log
    transactions = []
    running_balance = Decimal('0.00')
    
    for item in order_items:
        running_balance += item.seller_earnings
        transactions.append({
            'date': item.order.created_at,
            'description': f"Order #{item.order.id} – {item.product.name}",
            'in': item.seller_earnings,
            'out': Decimal('0.00'),
            'balance': running_balance,
        })
    
    # Get refunds
    refunds = Refund.objects.filter(
        seller=seller,
        created_at__gte=start_datetime,
        created_at__lte=end_datetime,
        status=Refund.STATUS_SUCCEEDED
    ).select_related('order', 'order_item', 'order_item__product')
    
    for refund in refunds:
        running_balance -= refund.amount
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        else:
            product_name = 'Full Order'
        
        transactions.append({
            'date': refund.created_at,
            'description': f"Refund #{refund.id} – Order #{refund.order.id} – {product_name}",
            'in': Decimal('0.00'),
            'out': refund.amount,
            'balance': running_balance,
        })
    
    # Sort by date
    transactions.sort(key=lambda x: x['date'])
    
    # Write CSV
    writer.writerow(['Earnings Statement', f'{start_dt} to {end_dt}'])
    writer.writerow([])
    writer.writerow(['Date', 'Description', 'In', 'Out', 'Balance'])
    
    for t in transactions:
        writer.writerow([
            t['date'].strftime('%Y-%m-%d %H:%M:%S'),
            t['description'],
            f"{t['in']:.2f}" if t['in'] > 0 else '',
            f"{t['out']:.2f}" if t['out'] > 0 else '',
            f"{t['balance']:.2f}",
        ])
    
    return response


# ==================== Excel, JSON, and PDF Export Functions ====================
# Note: These functions require additional packages (openpyxl, reportlab)
# They will fall back to CSV if packages are not installed

def export_orders_excel(seller, start_date, end_date, status_filter, product_filter):
    """Export seller's orders to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return export_orders_csv(seller, start_date, end_date, status_filter, product_filter)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['Order ID', 'Order Date', 'Customer Email', 'Order Status', 'Product Name', 'Quantity', 'Unit Price', 'Line Total', 'Platform Fee', 'Seller Earnings', 'Shipping Address', 'Tracking Number', 'Shipping Carrier']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    order_items = OrderItem.objects.filter(seller=seller).select_related('order', 'order__user', 'product')
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            order_items = order_items.filter(order__created_at__gte=start_dt)
        except:
            pass
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), datetime.max.time()))
            order_items = order_items.filter(order__created_at__lte=end_dt)
        except:
            pass
    if status_filter:
        order_items = order_items.filter(order__status=status_filter)
    if product_filter:
        order_items = order_items.filter(product_id=product_filter)
    for item in order_items.order_by('-order__created_at'):
        order = item.order
        product = item.product
        shipping_address = ""
        if order.is_pickup and order.pickup_location:
            shipping_address = f"PICKUP: {order.pickup_location.name}, {order.pickup_location.address1}, {order.pickup_location.city}"
        else:
            parts = []
            if order.ship_name:
                parts.append(order.ship_name)
            if order.ship_address1:
                parts.append(order.ship_address1)
            if order.ship_city:
                parts.append(order.ship_city)
            if order.ship_province:
                parts.append(order.ship_province)
            if order.ship_postal_code:
                parts.append(order.ship_postal_code)
            shipping_address = ", ".join(parts)
        ws.append([order.id, order.created_at.strftime('%Y-%m-%d %H:%M:%S'), order.user.email if order.user else 'Guest', order.get_status_display(), product.name, item.quantity, float(item.price), float(item.line_total), float(item.platform_fee), float(item.seller_earnings), shipping_address, order.tracking_number or '', order.get_shipping_carrier_display() if order.shipping_carrier else ''])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"orders_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_products_excel(seller, status_filter):
    """Export seller's products to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return export_products_csv(seller, status_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['Product ID', 'Name', 'Category', 'Price', 'Quantity in Stock', 'Is Active', 'Is Digital', 'Is Service', 'Is Featured', 'Created At', 'Updated At']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    products = Product.objects.filter(seller=seller).select_related('category')
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    for product in products.order_by('-created_at'):
        ws.append([product.id, product.name, product.category.name if product.category else '', float(product.price), product.quantity_in_stock, 'Yes' if product.is_active else 'No', 'Yes' if product.is_digital else 'No', 'Yes' if product.is_service else 'No', 'Yes' if product.is_featured else 'No', product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '', product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else ''])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_refunds_excel(seller, start_date, end_date, status_filter):
    """Export seller's refunds to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return export_refunds_csv(seller, start_date, end_date, status_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Refunds"
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['Refund ID', 'Order ID', 'Product Name', 'Amount', 'Reason', 'Status', 'Created By', 'Created At', 'Stripe Refund ID']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    refunds = Refund.objects.filter(seller=seller).select_related('order', 'order_item', 'order_item__product', 'created_by')
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            refunds = refunds.filter(created_at__gte=start_dt)
        except:
            pass
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), datetime.max.time()))
            refunds = refunds.filter(created_at__lte=end_dt)
        except:
            pass
    if status_filter:
        refunds = refunds.filter(status=status_filter)
    for refund in refunds.order_by('-created_at'):
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        elif refund.order_item:
            product_name = 'N/A'
        else:
            product_name = 'Full Order Refund'
        ws.append([refund.id, refund.order.id, product_name, float(refund.amount), refund.reason or '', refund.get_status_display(), refund.created_by.email if refund.created_by else '', refund.created_at.strftime('%Y-%m-%d %H:%M:%S'), refund.stripe_refund_id or ''])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"refunds_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_statement_excel(seller, start_date, end_date):
    """Export seller's earnings statement to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return export_statement_csv(seller, start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    now = timezone.now()
    today = now.date()
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            start_dt = date(today.year, today.month, 1)
            end_dt = today
    else:
        start_dt = date(today.year, today.month, 1)
        end_dt = today
    start_datetime = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_dt, datetime.max.time()))
    order_items = OrderItem.objects.filter(seller=seller, order__created_at__gte=start_datetime, order__created_at__lte=end_datetime).select_related('order', 'product', 'order__user').order_by('order__created_at')
    transactions = []
    running_balance = Decimal('0.00')
    for item in order_items:
        running_balance += item.seller_earnings
        transactions.append({'date': item.order.created_at, 'description': f"Order #{item.order.id} – {item.product.name}", 'in': item.seller_earnings, 'out': Decimal('0.00'), 'balance': running_balance})
    refunds = Refund.objects.filter(seller=seller, created_at__gte=start_datetime, created_at__lte=end_datetime, status=Refund.STATUS_SUCCEEDED).select_related('order', 'order_item', 'order_item__product')
    for refund in refunds:
        running_balance -= refund.amount
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        else:
            product_name = 'Full Order'
        transactions.append({'date': refund.created_at, 'description': f"Refund #{refund.id} – Order #{refund.order.id} – {product_name}", 'in': Decimal('0.00'), 'out': refund.amount, 'balance': running_balance})
    transactions.sort(key=lambda x: x['date'])
    ws.append(['Earnings Statement', f'{start_dt} to {end_dt}'])
    ws.append([])
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['Date', 'Description', 'In', 'Out', 'Balance']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for t in transactions:
        ws.append([t['date'].strftime('%Y-%m-%d %H:%M:%S'), t['description'], float(t['in']) if t['in'] > 0 else '', float(t['out']) if t['out'] > 0 else '', float(t['balance'])])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"statement_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_orders_json(seller, start_date, end_date, status_filter, product_filter):
    """Export seller's orders to JSON"""
    order_items = OrderItem.objects.filter(seller=seller).select_related('order', 'order__user', 'product')
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            order_items = order_items.filter(order__created_at__gte=start_dt)
        except:
            pass
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), datetime.max.time()))
            order_items = order_items.filter(order__created_at__lte=end_dt)
        except:
            pass
    if status_filter:
        order_items = order_items.filter(order__status=status_filter)
    if product_filter:
        order_items = order_items.filter(product_id=product_filter)
    data = []
    for item in order_items.order_by('-order__created_at'):
        order = item.order
        product = item.product
        shipping_address = ""
        if order.is_pickup and order.pickup_location:
            shipping_address = f"PICKUP: {order.pickup_location.name}, {order.pickup_location.address1}, {order.pickup_location.city}"
        else:
            parts = []
            if order.ship_name:
                parts.append(order.ship_name)
            if order.ship_address1:
                parts.append(order.ship_address1)
            if order.ship_city:
                parts.append(order.ship_city)
            if order.ship_province:
                parts.append(order.ship_province)
            if order.ship_postal_code:
                parts.append(order.ship_postal_code)
            shipping_address = ", ".join(parts)
        data.append({'order_id': order.id, 'order_date': order.created_at.isoformat(), 'customer_email': order.user.email if order.user else 'Guest', 'order_status': order.get_status_display(), 'product_name': product.name, 'quantity': item.quantity, 'unit_price': str(item.price), 'line_total': str(item.line_total), 'platform_fee': str(item.platform_fee), 'seller_earnings': str(item.seller_earnings), 'shipping_address': shipping_address, 'tracking_number': order.tracking_number or '', 'shipping_carrier': order.get_shipping_carrier_display() if order.shipping_carrier else ''})
    response = JsonResponse({'orders': data}, json_dumps_params={'indent': 2})
    filename = f"orders_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_products_json(seller, status_filter):
    """Export seller's products to JSON"""
    products = Product.objects.filter(seller=seller).select_related('category')
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    data = []
    for product in products.order_by('-created_at'):
        data.append({'product_id': product.id, 'name': product.name, 'category': product.category.name if product.category else '', 'price': str(product.price), 'quantity_in_stock': product.quantity_in_stock, 'is_active': product.is_active, 'is_digital': product.is_digital, 'is_service': product.is_service, 'is_featured': product.is_featured, 'created_at': product.created_at.isoformat() if product.created_at else None, 'updated_at': product.updated_at.isoformat() if product.updated_at else None})
    response = JsonResponse({'products': data}, json_dumps_params={'indent': 2})
    filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_refunds_json(seller, start_date, end_date, status_filter):
    """Export seller's refunds to JSON"""
    refunds = Refund.objects.filter(seller=seller).select_related('order', 'order_item', 'order_item__product', 'created_by')
    if start_date:
        try:
            start_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            refunds = refunds.filter(created_at__gte=start_dt)
        except:
            pass
    if end_date:
        try:
            end_dt = timezone.make_aware(datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), datetime.max.time()))
            refunds = refunds.filter(created_at__lte=end_dt)
        except:
            pass
    if status_filter:
        refunds = refunds.filter(status=status_filter)
    data = []
    for refund in refunds.order_by('-created_at'):
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        elif refund.order_item:
            product_name = 'N/A'
        else:
            product_name = 'Full Order Refund'
        data.append({'refund_id': refund.id, 'order_id': refund.order.id, 'product_name': product_name, 'amount': str(refund.amount), 'reason': refund.reason or '', 'status': refund.get_status_display(), 'created_by': refund.created_by.email if refund.created_by else '', 'created_at': refund.created_at.isoformat(), 'stripe_refund_id': refund.stripe_refund_id or ''})
    response = JsonResponse({'refunds': data}, json_dumps_params={'indent': 2})
    filename = f"refunds_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_statement_json(seller, start_date, end_date):
    """Export seller's earnings statement to JSON"""
    now = timezone.now()
    today = now.date()
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            start_dt = date(today.year, today.month, 1)
            end_dt = today
    else:
        start_dt = date(today.year, today.month, 1)
        end_dt = today
    start_datetime = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_dt, datetime.max.time()))
    order_items = OrderItem.objects.filter(seller=seller, order__created_at__gte=start_datetime, order__created_at__lte=end_datetime).select_related('order', 'product', 'order__user').order_by('order__created_at')
    transactions = []
    running_balance = Decimal('0.00')
    for item in order_items:
        running_balance += item.seller_earnings
        transactions.append({'date': item.order.created_at.isoformat(), 'description': f"Order #{item.order.id} – {item.product.name}", 'in': str(item.seller_earnings), 'out': '0.00', 'balance': str(running_balance)})
    refunds = Refund.objects.filter(seller=seller, created_at__gte=start_datetime, created_at__lte=end_datetime, status=Refund.STATUS_SUCCEEDED).select_related('order', 'order_item', 'order_item__product')
    for refund in refunds:
        running_balance -= refund.amount
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        else:
            product_name = 'Full Order'
        transactions.append({'date': refund.created_at.isoformat(), 'description': f"Refund #{refund.id} – Order #{refund.order.id} – {product_name}", 'in': '0.00', 'out': str(refund.amount), 'balance': str(running_balance)})
    transactions.sort(key=lambda x: x['date'])
    response = JsonResponse({'statement': {'period': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()}, 'transactions': transactions}}, json_dumps_params={'indent': 2})
    filename = f"statement_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_statement_pdf(seller, start_date, end_date):
    """Export seller's earnings statement to PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
    except ImportError:
        return export_statement_csv(seller, start_date, end_date)
    now = timezone.now()
    today = now.date()
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            start_dt = date(today.year, today.month, 1)
            end_dt = today
    else:
        start_dt = date(today.year, today.month, 1)
        end_dt = today
    start_datetime = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_dt, datetime.max.time()))
    order_items = OrderItem.objects.filter(seller=seller, order__created_at__gte=start_datetime, order__created_at__lte=end_datetime).select_related('order', 'product', 'order__user').order_by('order__created_at')
    transactions = []
    running_balance = Decimal('0.00')
    total_in = Decimal('0.00')
    total_out = Decimal('0.00')
    for item in order_items:
        running_balance += item.seller_earnings
        total_in += item.seller_earnings
        transactions.append({'date': item.order.created_at, 'description': f"Order #{item.order.id} – {item.product.name}", 'in': item.seller_earnings, 'out': Decimal('0.00'), 'balance': running_balance})
    refunds = Refund.objects.filter(seller=seller, created_at__gte=start_datetime, created_at__lte=end_datetime, status=Refund.STATUS_SUCCEEDED).select_related('order', 'order_item', 'order_item__product')
    for refund in refunds:
        running_balance -= refund.amount
        total_out += refund.amount
        product_name = ''
        if refund.order_item and refund.order_item.product:
            product_name = refund.order_item.product.name
        else:
            product_name = 'Full Order'
        transactions.append({'date': refund.created_at, 'description': f"Refund #{refund.id} – Order #{refund.order.id} – {product_name}", 'in': Decimal('0.00'), 'out': refund.amount, 'balance': running_balance})
    transactions.sort(key=lambda x: x['date'])
    response = HttpResponse(content_type='application/pdf')
    filename = f"statement_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#2d6a4f'), spaceAfter=30)
    story.append(Paragraph("Earnings Statement", title_style))
    story.append(Paragraph(f"Period: {start_dt.strftime('%B %d, %Y')} to {end_dt.strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Paragraph(f"Seller: {seller.display_name or seller.user.email}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    summary_data = [['Total In', f"${total_in:.2f}"], ['Total Out', f"${total_out:.2f}"], ['Net Change', f"${total_in - total_out:.2f}"]]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey), ('TEXTCOLOR', (0, 0), (-1, -1), colors.black), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 12), ('BOTTOMPADDING', (0, 0), (-1, -1), 12), ('TOPPADDING', (0, 0), (-1, -1), 12)]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    table_data = [['Date', 'Description', 'In', 'Out', 'Balance']]
    for t in transactions:
        table_data.append([t['date'].strftime('%Y-%m-%d'), t['description'], f"${t['in']:.2f}" if t['in'] > 0 else '', f"${t['out']:.2f}" if t['out'] > 0 else '', f"${t['balance']:.2f}"])
    transactions_table = Table(table_data, colWidths=[1*inch, 3.5*inch, 1*inch, 1*inch, 1*inch])
    transactions_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d6a4f')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('FONTSIZE', (0, 1), (-1, -1), 9)]))
    story.append(transactions_table)
    doc.build(story)
    return response


@seller_required
def membership_plans_list(request):
    """List all membership plans for the seller"""
    seller = request.user.seller
    plans = SellerMembershipPlan.objects.filter(seller=seller).order_by('display_order', 'name')
    
    return render(request, 'sellers/membership_plans_list.html', {
        'plans': plans,
        'seller': seller,
    })


@seller_required
def membership_plan_add(request):
    """Add a new membership plan"""
    seller = request.user.seller
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        slug = request.POST.get('slug', '').strip()
        price = request.POST.get('price', '0.00')
        description = request.POST.get('description', '')
        details = request.POST.get('details', '')
        is_active = request.POST.get('is_active') == 'on'
        display_order = request.POST.get('display_order', '0')
        
        # Validation
        if not name:
            messages.error(request, 'Plan name is required.')
            return redirect('sellers:membership_plan_add')
        
        if not slug:
            # Auto-generate slug from name
            slug = name.lower().replace(' ', '-').replace('_', '-')
            # Remove special characters
            slug = ''.join(c for c in slug if c.isalnum() or c == '-')
        
        # Check if slug already exists for this seller
        if SellerMembershipPlan.objects.filter(seller=seller, slug=slug).exists():
            messages.error(request, f'A plan with slug "{slug}" already exists. Please choose a different slug.')
            return redirect('sellers:membership_plan_add')
        
        try:
            price_decimal = Decimal(str(price))
            display_order_int = int(display_order) if display_order else 0
        except (ValueError, TypeError):
            messages.error(request, 'Invalid price or display order.')
            return redirect('sellers:membership_plan_add')
        
        # Create the plan
        plan = SellerMembershipPlan.objects.create(
            seller=seller,
            name=name,
            slug=slug,
            price=price_decimal,
            description=description,
            details=details,
            is_active=is_active,
            display_order=display_order_int,
        )
        
        messages.success(request, f'Membership plan "{plan.name}" has been created successfully!')
        return redirect('sellers:membership_plans_list')
    
    return render(request, 'sellers/membership_plan_form.html', {
        'seller': seller,
        'action': 'Add',
    })


@seller_required
def membership_plan_edit(request, plan_id):
    """Edit an existing membership plan"""
    seller = request.user.seller
    plan = get_object_or_404(SellerMembershipPlan, id=plan_id, seller=seller)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        slug = request.POST.get('slug', '').strip()
        price = request.POST.get('price', '0.00')
        description = request.POST.get('description', '')
        details = request.POST.get('details', '')
        is_active = request.POST.get('is_active') == 'on'
        display_order = request.POST.get('display_order', '0')
        
        # Validation
        if not name:
            messages.error(request, 'Plan name is required.')
            return redirect('sellers:membership_plan_edit', plan_id=plan_id)
        
        if not slug:
            # Auto-generate slug from name
            slug = name.lower().replace(' ', '-').replace('_', '-')
            # Remove special characters
            slug = ''.join(c for c in slug if c.isalnum() or c == '-')
        
        # Check if slug already exists for this seller (excluding current plan)
        if SellerMembershipPlan.objects.filter(seller=seller, slug=slug).exclude(id=plan_id).exists():
            messages.error(request, f'A plan with slug "{slug}" already exists. Please choose a different slug.')
            return redirect('sellers:membership_plan_edit', plan_id=plan_id)
        
        try:
            price_decimal = Decimal(str(price))
            display_order_int = int(display_order) if display_order else 0
        except (ValueError, TypeError):
            messages.error(request, 'Invalid price or display order.')
            return redirect('sellers:membership_plan_edit', plan_id=plan_id)
        
        # Check if plan has active members before deactivating
        if not is_active and plan.is_active and plan.has_active_members():
            active_count = plan.get_active_member_count()
            messages.warning(request, f'Cannot deactivate plan "{plan.name}" - it has {active_count} active member subscription(s).')
            return redirect('sellers:membership_plan_edit', plan_id=plan_id)
        
        # Update the plan
        plan.name = name
        plan.slug = slug
        plan.price = price_decimal
        plan.description = description
        plan.details = details
        plan.is_active = is_active
        plan.display_order = display_order_int
        plan.save()
        
        messages.success(request, f'Membership plan "{plan.name}" has been updated successfully!')
        return redirect('sellers:membership_plans_list')
    
    return render(request, 'sellers/membership_plan_form.html', {
        'seller': seller,
        'plan': plan,
        'action': 'Edit',
    })


@seller_required
def membership_plan_delete(request, plan_id):
    """Delete a membership plan"""
    seller = request.user.seller
    plan = get_object_or_404(SellerMembershipPlan, id=plan_id, seller=seller)
    
    if request.method == 'POST':
        # Check if plan has active members
        if plan.has_active_members():
            active_count = plan.get_active_member_count()
            messages.error(
                request,
                f'Cannot delete plan "{plan.name}" - it has {active_count} active member subscription(s). Please inactivate instead.'
            )
            return redirect('sellers:membership_plans_list')
        
        plan_name = plan.name
        plan.delete()
        messages.success(request, f'Membership plan "{plan_name}" has been deleted successfully!')
        return redirect('sellers:membership_plans_list')
    
    return render(request, 'sellers/membership_plan_confirm_delete.html', {
        'plan': plan,
    })


@seller_required
def membership_plan_toggle_active(request, plan_id):
    """Toggle active/inactive status of a membership plan"""
    seller = request.user.seller
    plan = get_object_or_404(SellerMembershipPlan, id=plan_id, seller=seller)
    
    if request.method == 'POST':
        # If trying to deactivate, check for active members
        if plan.is_active and plan.has_active_members():
            active_count = plan.get_active_member_count()
            messages.warning(
                request,
                f'Cannot deactivate plan "{plan.name}" - it has {active_count} active member subscription(s).'
            )
        else:
            plan.is_active = not plan.is_active
            plan.save()
            status = "activated" if plan.is_active else "deactivated"
            messages.success(request, f'Membership plan "{plan.name}" has been {status}.')
    
    return redirect('sellers:membership_plans_list')
